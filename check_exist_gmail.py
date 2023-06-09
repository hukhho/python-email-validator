import openpyxl
from validate_email_address import validate_email
from concurrent.futures import ThreadPoolExecutor, as_completed

def check_email_validity(email):
    if not isinstance(email, str):
        return "Invalid"
    is_valid = validate_email(email, verify=True)
    return "Live" if is_valid else "Die"

def process_row(email_cell):
    email = email_cell.value
    if email is not None and email != "":
        email_status = check_email_validity(email)
        return email_cell.row, email_status
    return None, None

# Read the input data
input_workbook = openpyxl.load_workbook('input.xlsx')
input_sheet = input_workbook.active
total_rows = input_sheet.max_row - 2

with ThreadPoolExecutor(max_workers=50) as executor:
    futures = [executor.submit(process_row, row[3]) for row in input_sheet.iter_rows(min_row=3, max_row=input_sheet.max_row, values_only=False)]
    
    completed_count = 0
    for future in as_completed(futures):
        row_number, result = future.result()
        if row_number is not None:
            input_sheet.cell(row=row_number, column=5).value = result
        completed_count += 1
        progress_percentage = (completed_count / total_rows) * 100
        print(f"Progress: {progress_percentage:.2f}%")

input_workbook.save('output.xlsx')
print("Email status added to output.xlsx")